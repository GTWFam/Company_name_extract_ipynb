from difflib import SequenceMatcher
import pandas as pd
import numpy as np
import json
import string

from openpyxl import load_workbook

leadFeeder = pd.read_excel('./sheets/May Leadlander Data.xlsx')
wb = load_workbook(filename='./sheets/dtex-2022-06-03-10-08-29[4].xlsx')
adDaptive = pd.read_excel(
    './sheets/dtex-2022-06-03-10-08-29[4].xlsx', sheet_name=wb.sheetnames[1])

wb_target = load_workbook(filename='sheets/Target Account List (3).xlsx')
target_lvl1 = pd.read_excel(
    './sheets/Target Account List (3).xlsx', sheet_name=wb_target.sheetnames[0])
target_lvl2 = pd.read_excel(
    './sheets/Target Account List (3).xlsx', sheet_name=wb_target.sheetnames[1])
target_lvl3 = pd.read_excel(
    './sheets/Target Account List (3).xlsx', sheet_name=wb_target.sheetnames[2])

target_lvl1_companies = target_lvl1["Level 1"].to_numpy()
target_lvl2_companies = target_lvl2["Level 2"].to_numpy()
target_lvl3_companies = target_lvl3["Level 3"].to_numpy()

adDaptive.columns = adDaptive.iloc[6]
adDaptive = adDaptive.iloc[7:, :]
adDaptive.reset_index(drop=True, inplace=True)

ad_companies = adDaptive["Business Name"].to_numpy()
lf_companies = leadFeeder["Company"].to_numpy()


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


def clean_name(name):
    name = name.lower()
    name = name.translate(
        str.maketrans('', '', string.punctuation))
    name = name.strip()
    name = name.replace("  ", ' ')
    name = name.replace('inc', '')
    name = name.replace("llc", '')
    return name


def compute_cross_helper(comp_list, target, cross, key, list_type):
    remain = []
    for comp in comp_list:
        comp_name = clean_name(comp)
        sim = similar(target, comp_name)
        if (sim > 0.8 and comp.lower().startswith(target.split(" ")[0] + " ")) or sim > 0.85:
            cross[key][list_type] += [comp]
        else:
            remain.append(comp)
    return cross, remain


def cross_helper_detail(comp_list, target, cross, key, list_type):
    remain = []
    for comp in comp_list:
        comp_name = clean_name(comp)
        sim = similar(target, comp_name)
        if sim > 0.4 and target + " " in comp_name and comp_name.startswith(target):
            cross[key][list_type] += [comp]
        else:
            remain.append(comp)
    return cross, remain


def compute_cross(target, adDaptive, leadFeeder, name):
    cross_list = {}
    [cross_list.setdefault(x, {"adDaptive": [], "LeadFeeder": []})
     for x in target]

    for comp in target:
        comp_name = comp.lower()
        cross_list, adDaptive = compute_cross_helper(
            adDaptive, comp_name, cross_list, comp, "adDaptive")
        cross_list, leadFeeder = compute_cross_helper(
            leadFeeder, comp_name, cross_list, comp, "LeadFeeder")

    for comp in target:
        comp_name = comp.lower()
        cross_list, adDaptive = cross_helper_detail(
            adDaptive, comp_name, cross_list, comp, "adDaptive")
        cross_list, leadFeeder = cross_helper_detail(
            leadFeeder, comp_name, cross_list, comp, "LeadFeeder")

    cross_list["remain"] = {
        "adDaptive": adDaptive, "LeadFeeder": leadFeeder}
    with open('./' + name + '.json', 'w') as f:
        json.dump(cross_list, f, indent=4)


compute_cross(target_lvl1_companies, ad_companies, lf_companies, "lvl1")
