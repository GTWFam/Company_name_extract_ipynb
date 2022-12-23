import pandas as pd
import numpy as np
import json
import string

from openpyxl import load_workbook

leadFeeder = pd.read_excel('./May Leadlander Data.xlsx')
wb = load_workbook(filename='./dtex-2022-06-03-10-08-29[4].xlsx')
adDaptive = pd.read_excel(
    './dtex-2022-06-03-10-08-29[4].xlsx', sheet_name=wb.sheetnames[1])

adDaptive.columns = adDaptive.iloc[6]
adDaptive = adDaptive.iloc[7:, :]
adDaptive.reset_index(drop=True, inplace=True)

ad_companies = adDaptive["Business Name"].to_numpy()
lf_companies = leadFeeder["Company"].to_numpy()

common_words = ['service', 'united', 'bank', 'agency', 'postal', 'administration', 'federal', 'insurance', 'inc', 'co', 'llc', 'ltd' 'network', 'financial', 'mutual', 'state', 'corporation', 'union', 'company', 'corp', 'incorporated', 'health', 'and', 'group', 'america', 'american', 'clinic', 'LLP', 'international', 'district', 'capital', 'living', 'loans', 'healthcare', 'john', 'mortgage', 'electric', 'communications', 'carrier', 'systems', 'services', 'management', 'california',
                'trust', 'national', 'on', 'auto', 'consulting', 'foundation', 'the', 'general', 'cowen', 'system', 'entertainment', 'nga', 'technologies', 'university', 'fitness', 'us', 'johnson', 'tech', 'research', 'partners', 'software', 'enterprises', 'blue', 'commission', 'properties', 'networks', 'foods', 'mae', 'solutions', 'chase', 'associates', 'farms', 'designs', 'materials', 'automation', 'training', 'greif', 'partnership', 'advisors', 'assoc', 'payments', 'air', 'technology', 'warehouse', 'states', 'third', 'family', 'Stanley', 'Companies', 'Credit', 'First', 'Medical', 'Express', 'Gas', 'Global', 'Pacific', 'Universal']


def similar_name(cp_name, companies):
    name = cp_name.lower()
    name = name.translate(str.maketrans(
        '', '', string.punctuation))
    name = name.strip()
    name = name.replace("  ", ' ')
    words = name.split(' ')
    for company in companies:
        comp = company.lower()
        for word in words:
            if word in comp.split(' ') and word not in common_words:
                return True, cp_name, company
    return False, '', ''


ad_parent = {}
lf_parent = {}

for comp in ad_companies:
    flag, similar_word, company = similar_name(comp, ad_parent.keys())
    if flag:
        ad_parent[company] += [similar_word]
    else:
        if ad_parent.get(comp) is None:
            ad_parent[comp] = [comp]
        else:
            ad_parent[comp] += [comp]
with open('./test.json', 'w') as f:
    json.dump(ad_parent, f, indent=4)
